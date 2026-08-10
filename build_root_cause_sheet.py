#!/usr/bin/env python3
"""Build a root-cause-analysis workbook from data/issue_rca.json.

Two tabs:
  - "Root Cause Themes": the ~9 cross-cutting root-cause themes that the 15
    issue buckets actually reduce to, each with a recommended automated fix
    and an owner - this is the actionable list to build a remediation
    workflow against, since most individual issue codes are downstream
    symptoms of the same handful of underlying causes.
  - "Issue x Root Cause Detail": one row per root-cause finding (the content
    of data/issue_rca.json flattened), tagged with which theme it belongs to.

Theme assignment is a manual judgment call made while writing the RCA
(root_causes were derived from reading real transcripts/evidence per bucket,
see build_issue_bucket_samples.py) - it's encoded here as a static map rather
than re-derived, so re-running this script after refreshing issue_rca.json
requires updating THEME_ASSIGNMENTS to match any new/changed bullets.

Usage:
    python3 build_root_cause_sheet.py
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
RCA_PATH = os.path.join(HERE, "data", "issue_rca.json")
SAMPLES_PATH = os.path.join(HERE, "data", "issue_bucket_samples.json")
OUT_PATH = os.path.join(HERE, "reports", "root_cause_analysis.xlsx")

# One-line plain-English definition of what each issue code means, independent
# of the root-cause narrative - so the sheet is readable without also having
# the dashboard open.
ISSUE_MEANINGS = {
    "high_word_error_rate": "Speech-to-text transcript error rate vs. reference exceeded threshold - flags likely audio/ASR quality problems on the call.",
    "speaking_pace_too_fast": "Agent's words-per-minute on one or more turns exceeded the fast-pace threshold.",
    "call_flow_not_followed": "Agent skipped or deviated from the dealership's expected scripted call flow (e.g. skipped the mandated opener or a required qualifying step).",
    "high_response_latency": "Gap between the customer finishing speaking and the agent responding exceeded the latency threshold on one or more turns.",
    "opening_greeting_missing": "Agent never delivered the required opening greeting/self-identification at the start of the call.",
    "not_interested": "Customer explicitly declined to engage with the sales pitch.",
    "speaking_pace_too_slow": "Agent's words-per-minute on one or more turns fell below the slow-pace threshold.",
    "not_responding": "Customer went silent despite the agent's engagement attempts.",
    "agent_silence": "Agent left a gap longer than the silence threshold without responding.",
    "ivr_detected": "Call was flagged as answered by an automated IVR/recorded message rather than a person.",
    "user_not_available": "Someone other than the intended contact answered and said that person wasn't available.",
    "wrong_number": "The person who answered confirmed they are not the intended contact.",
    "irrelevant_intent": "Caller's request was unrelated to the dealership's sales offering.",
    "voicemail_detected": "Call was routed to voicemail.",
    "already_addressed": "Customer indicated their need was already resolved (e.g. already purchased) before the call.",
}

THEMES = {
    "NONHUMAN": {
        "short": "Non-human pickup mishandled",
        "name": "Non-human pickup mishandled (voicemail / IVR / carrier announcement / screening bot)",
        "category": "Detection & routing fix",
        "action": "Add an answering-machine/IVR/screener classifier that runs before the conversational flow starts. Branch to a dedicated short flow per pickup type (leave one concise VM message and hang up; answer a screener's questions then stop; disengage immediately from a robocall/hold-queue) instead of running the live-call script. Suppress WPM/latency/WER/greeting/call-flow flags on calls classified this way - those metrics don't mean anything against a machine.",
        "owner": "Eng - Voice platform / call routing",
        "priority": "High",
    },
    "OPENERTRIGGER": {
        "short": "Opener doesn't fire when callee speaks first",
        "name": "Scripted opener doesn't fire when the callee speaks first",
        "category": "Agent script/logic fix",
        "action": "Redesign the opening turn so the agent always leads with greeting + self-ID regardless of what the callee says first (e.g. 'Speaking.', 'Hello?', 'Is this X?'). Treat a bare acknowledgment as a cue to proceed into the full greeting, not as a completed exchange.",
        "owner": "Eng - Conversation design",
        "priority": "High",
    },
    "EXITLOGIC": {
        "short": "No graceful exit on rejection signals",
        "name": "Missing graceful-exit branch on a clear negative/terminal signal",
        "category": "Agent script/logic fix",
        "action": "Add explicit intent classifiers for rejection / already-purchased / wrong-person / unavailable / off-scope-request that short-circuit straight to a one-line close-out and call termination, instead of falling through to the default script tree or re-asking the same question.",
        "owner": "Eng - Conversation design",
        "priority": "High",
    },
    "LISTDATA": {
        "short": "Stale CRM / dial-list data",
        "name": "Stale or incorrect CRM / dial-list data",
        "category": "CRM data feedback loop",
        "action": "Feed call verdicts back into the dialer/CRM automatically: when a call resolves as wrong_number, already_addressed, or an explicit do-not-call request, auto-flag that contact/number so it's excluded from future campaigns. Today the agent sometimes promises removal in-call but the list isn't actually updated.",
        "owner": "RevOps - list hygiene / CRM integration",
        "priority": "High",
    },
    "SCRIPTPACING": {
        "short": "Script crams too much into one turn",
        "name": "Script crams multiple beats into one breath / one turn runs too long",
        "category": "Agent script/logic fix",
        "action": "Split compound opener and wrap-up lines (name-confirm + self-ID + dealership + purpose; or decline-ack + courtesy line) into shorter turns with a natural pause between beats.",
        "owner": "Eng - Conversation design",
        "priority": "Medium",
    },
    "DETECTORBUG": {
        "short": "Detector calibration false positives",
        "name": "Detector/metric calibration false positive (the measurement is wrong, not the agent's behavior)",
        "category": "Detector calibration fix",
        "action": "Patch the scoring pipeline directly: exclude calls below a minimum transcript length from WER scoring, exclude digit-spelled-out utterances from WPM calculation, allowlist the agent's own mandatory disclosure phrase in IVR detection, and replace voicemail keyword-matching with a check for whether a complete message was actually left.",
        "owner": "Eng - Audit/scoring pipeline",
        "priority": "Medium",
    },
    "LATENCYFILLER": {
        "short": "No filler phrase during tool latency",
        "name": "Tool-call / lookup latency with no filler phrase to mask it",
        "category": "Agent script/logic fix",
        "action": "Insert a short spoken filler ('Let me check that for you, one moment') immediately before invoking any lookup/tool that can take multiple seconds, so the gap doesn't register as dead air or high latency.",
        "owner": "Eng - Conversation design",
        "priority": "Medium",
    },
    "AUDIOQUALITY": {
        "short": "Real audio/line quality issues",
        "name": "Real audio/line quality issues (echo, crosstalk, dropped connection)",
        "category": "Telephony infrastructure",
        "action": "Not fixable in the agent's script - track separately as a carrier/line-quality metric (echo cancellation, jitter, disconnect rate) rather than folding it into agent-performance issue codes.",
        "owner": "Eng - Telephony/infra",
        "priority": "Low",
    },
    "CUSTBEHAVIOR": {
        "short": "Customer behavior, not a system defect",
        "name": "Genuine customer behavior, not a system defect",
        "category": "Not actionable - reclassify/suppress",
        "action": "Auto-classify these as non-actionable at audit time (customer disengaged, declined the recording-consent disclosure, or was never actually engaged) so they stop consuming engineering/product review time as if they were agent failures.",
        "owner": "Eng - Audit/scoring pipeline",
        "priority": "Low",
    },
}

# (issue_code, root_cause_index) -> theme key. Index is 0-based position within
# that bucket's root_causes list in data/issue_rca.json.
THEME_ASSIGNMENTS = {
    ("high_word_error_rate", 0): "DETECTORBUG",
    ("high_word_error_rate", 1): "AUDIOQUALITY",
    ("high_word_error_rate", 2): "NONHUMAN",
    ("speaking_pace_too_fast", 0): "SCRIPTPACING",
    ("speaking_pace_too_fast", 1): "NONHUMAN",
    ("speaking_pace_too_fast", 2): "NONHUMAN",
    ("speaking_pace_too_fast", 3): "SCRIPTPACING",
    ("call_flow_not_followed", 0): "OPENERTRIGGER",
    ("call_flow_not_followed", 1): "NONHUMAN",
    ("call_flow_not_followed", 2): "EXITLOGIC",
    ("call_flow_not_followed", 3): "AUDIOQUALITY",
    ("high_response_latency", 0): "NONHUMAN",
    ("high_response_latency", 1): "OPENERTRIGGER",
    ("high_response_latency", 2): "LATENCYFILLER",
    ("opening_greeting_missing", 0): "OPENERTRIGGER",
    ("opening_greeting_missing", 1): "SCRIPTPACING",
    ("opening_greeting_missing", 2): "NONHUMAN",
    ("opening_greeting_missing", 3): "AUDIOQUALITY",
    ("not_interested", 0): "SCRIPTPACING",
    ("not_interested", 1): "LISTDATA",
    ("not_interested", 2): "CUSTBEHAVIOR",
    ("not_interested", 3): "LISTDATA",
    ("speaking_pace_too_slow", 0): "DETECTORBUG",
    ("speaking_pace_too_slow", 1): "DETECTORBUG",
    ("speaking_pace_too_slow", 2): "DETECTORBUG",
    ("speaking_pace_too_slow", 3): "LATENCYFILLER",
    ("not_responding", 0): "NONHUMAN",
    ("not_responding", 1): "NONHUMAN",
    ("not_responding", 2): "CUSTBEHAVIOR",
    ("not_responding", 3): "AUDIOQUALITY",
    ("agent_silence", 0): "LATENCYFILLER",
    ("agent_silence", 1): "OPENERTRIGGER",
    ("agent_silence", 2): "CUSTBEHAVIOR",
    ("ivr_detected", 0): "DETECTORBUG",
    ("ivr_detected", 1): "NONHUMAN",
    ("ivr_detected", 2): "NONHUMAN",
    ("user_not_available", 0): "NONHUMAN",
    ("user_not_available", 1): "EXITLOGIC",
    ("user_not_available", 2): "EXITLOGIC",
    ("wrong_number", 0): "LISTDATA",
    ("wrong_number", 1): "LISTDATA",
    ("wrong_number", 2): "EXITLOGIC",
    ("wrong_number", 3): "NONHUMAN",
    ("irrelevant_intent", 0): "LISTDATA",
    ("irrelevant_intent", 1): "LISTDATA",
    ("irrelevant_intent", 2): "EXITLOGIC",
    ("irrelevant_intent", 3): "EXITLOGIC",
    ("voicemail_detected", 0): "NONHUMAN",
    ("voicemail_detected", 1): "NONHUMAN",
    ("voicemail_detected", 2): "DETECTORBUG",
    ("already_addressed", 0): "LISTDATA",
    ("already_addressed", 1): "EXITLOGIC",
    ("already_addressed", 2): "LISTDATA",
}

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
PRIORITY_FILL = {
    "High": PatternFill(start_color="FCE2E2", end_color="FCE2E2", fill_type="solid"),
    "Medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "Low": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
}


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_themes_sheet(wb, buckets):
    ws = wb.active
    ws.title = "Root Cause Themes"
    headers = [
        "Theme",
        "Category",
        "Issue codes it shows up in (code - what it means)",
        "Red-alert calls in those buckets (total, not per-theme - see note)",
        "Recommended automated workflow action",
        "Owner",
        "Priority",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    codes_by_theme = {}
    counts_by_theme = {}
    for b in buckets:
        code = b["code"]
        total = b["total_count"]
        for idx in range(len(b["root_causes"])):
            theme = THEME_ASSIGNMENTS.get((code, idx))
            if not theme:
                continue
            codes_by_theme.setdefault(theme, set()).add(code)
            counts_by_theme.setdefault(theme, {})[code] = total

    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    theme_items = sorted(THEMES.items(), key=lambda kv: priority_order.get(kv[1]["priority"], 9))

    for key, t in theme_items:
        codes = sorted(codes_by_theme.get(key, []))
        codes_explained = "; ".join(f"{c} - {ISSUE_MEANINGS.get(c, '')}" for c in codes)
        total = sum(counts_by_theme.get(key, {}).values())
        row = [
            t["name"],
            t["category"],
            codes_explained,
            total,
            t["action"],
            t["owner"],
            t["priority"],
        ]
        ws.append(row)
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
        fill = PRIORITY_FILL.get(t["priority"])
        if fill:
            ws.cell(row=r, column=7).fill = fill

    autosize(ws, [42, 22, 65, 14, 60, 26, 12])

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1,
             value="Note: the call count is the total size of every bucket this theme touches, not a precise per-theme "
                   "count - themes were assigned by reading a sample of calls per bucket, not by tagging every call. "
                   "Treat it as a rough scale indicator, not an exact figure.").alignment = WRAP


def load_evidence_lookup():
    """(issue_code, call_id) -> the specific per-call evidence text that made
    SuperBryn's own audit flag that issue on that call - i.e. what exactly
    failed on that one call, as opposed to the aggregated bucket-level pattern."""
    if not os.path.exists(SAMPLES_PATH):
        return {}
    with open(SAMPLES_PATH) as f:
        samples = json.load(f)
    lookup = {}
    for b in samples:
        for c in b["sample_calls"]:
            lookup[(b["code"], c["id"])] = c.get("evidence")
    return lookup


def build_detail_sheet(wb, buckets):
    ws = wb.create_sheet("Issue x Root Cause Detail")
    headers = [
        "Issue code",
        "Issue meaning",
        "Lane",
        "Bucket total (red-alert calls)",
        "Root cause pattern",
        "Theme",
        "Call ID",
        "What exactly failed on this call",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    evidence_lookup = load_evidence_lookup()

    for b in buckets:
        for idx, rc in enumerate(b["root_causes"]):
            theme_key = THEME_ASSIGNMENTS.get((b["code"], idx))
            theme_name = THEMES[theme_key]["name"] if theme_key else ""
            for call_id in rc["example_call_ids"]:
                evidence = evidence_lookup.get((b["code"], call_id), "")
                ws.append([
                    b["code"],
                    ISSUE_MEANINGS.get(b["code"], b["title"]),
                    b["lane"],
                    b["total_count"],
                    rc["summary"],
                    theme_name,
                    call_id,
                    evidence,
                ])
                r = ws.max_row
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).alignment = WRAP

    autosize(ws, [24, 46, 16, 12, 60, 42, 38, 55])


def build_theme_export(buckets):
    """The same theme/owner/priority/action rollup that goes into the 'Root Cause Themes'
    xlsx tab, as JSON - so the live dashboard (server.py) can show it without re-deriving
    anything or duplicating THEMES/THEME_ASSIGNMENTS in a second place."""
    codes_by_theme = {}
    counts_by_theme = {}
    issue_theme_by_code = {}
    for b in buckets:
        code = b["code"]
        issue_theme_by_code[code] = []
        for idx in range(len(b["root_causes"])):
            theme = THEME_ASSIGNMENTS.get((code, idx))
            issue_theme_by_code[code].append(theme)
            if not theme:
                continue
            codes_by_theme.setdefault(theme, set()).add(code)
            counts_by_theme.setdefault(theme, {})[code] = b["total_count"]

    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    themes = []
    for key, t in sorted(THEMES.items(), key=lambda kv: priority_order.get(kv[1]["priority"], 9)):
        themes.append({
            "key": key,
            "short": t["short"],
            "name": t["name"],
            "category": t["category"],
            "action": t["action"],
            "owner": t["owner"],
            "priority": t["priority"],
            "issue_codes": sorted(codes_by_theme.get(key, [])),
            "red_alert_calls": sum(counts_by_theme.get(key, {}).values()),
        })

    return {"themes": themes, "issue_theme_by_code": issue_theme_by_code}


def main():
    with open(RCA_PATH) as f:
        buckets = json.load(f)

    missing = [
        (b["code"], idx)
        for b in buckets
        for idx in range(len(b["root_causes"]))
        if (b["code"], idx) not in THEME_ASSIGNMENTS
    ]
    if missing:
        raise SystemExit(f"THEME_ASSIGNMENTS is missing entries for: {missing}")

    wb = Workbook()
    build_themes_sheet(wb, buckets)
    build_detail_sheet(wb, buckets)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(THEMES)} themes, {len(buckets)} issue buckets)")

    theme_export_path = os.path.join(HERE, "data", "root_cause_themes.json")
    with open(theme_export_path, "w") as f:
        json.dump(build_theme_export(buckets), f, indent=2)
    print(f"Wrote {theme_export_path}")


if __name__ == "__main__":
    main()
