#!/usr/bin/env python3
"""Build reports/quality_tracker_Aug2026.xlsx - a local mirror of the shared
Google Sheet quality tracker, with Ticket/Owner/Status/ETA columns per issue
row and one date column per audited day.

Day values are hardcoded here (pulled from the observability/red-alert data
for 1-11 Aug via server.py's compute_issues()-equivalent logic) rather than
recomputed, since this mirrors a manually-curated tracker - update DAYS/
SALES_INBOUND/SALES_OUTBOUND below when adding a new day.

Usage:
    python3 build_quality_tracker_sheet.py
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(HERE, "reports", "quality_tracker_Aug2026.xlsx")

DAYS = ["1 Aug", "2 Aug", "3 Aug", "4 Aug", "5 Aug", "6 Aug", "7 Aug", "8 Aug", "9 Aug", "10 Aug", "11 Aug"]

JIRA_ROWS = [
    ("https://spyne.atlassian.net/browse/RETCONVAI-4406", "", "Piyush",
     "Already rolled out for IB agents/ in pipeline for OB Waiting for livekit to increase concurrency"),
    ("https://spyne.atlassian.net/browse/RETCONVAI-4407", "", "Bhaskar", "In progress"),
    ("https://spyne.atlassian.net/browse/RETCONVAI-4408", "", "Bhaskar",
     "already fixed and deployed on production for new service prompt"),
]

# Each issue gets a count row plus a "% of red alerts" row directly beneath it - that
# issue's count as a percentage of the day's confirmed red alerts (TP+FN), not of audited
# calls. An issue can also be flagged on non-red (FP/TN) audited calls, so this can exceed
# 100%; "-" marks days with zero red alerts, where the ratio is undefined rather than 0%.
INBOUND_RED = [49, 15, 85, 41, 75, 42, 70, 39, 12, 97, 68]
OUTBOUND_RED = [81, 1, 117, 29, 139, 37, 109, 56, 0, 117, 82]


def issue_rows(title, counts, red):
    pct = [f"{round(c / r * 100)}%" if r else "-" for c, r in zip(counts, red)]
    return [(title, counts), ("    % of red alerts", pct)]


SALES_INBOUND = [
    ("Total calls", [132, 34, 289, 283, 266, 161, 239, 115, 30, 290, 252]),
    ("Audited calls", [70, 20, 119, 58, 123, 62, 96, 58, 16, 126, 105]),
    ("% Red Alerts", ["70%", "75%", "71%", "71%", "61%", "68%", "73%", "67%", "75%", "77%", "65%"]),
    *issue_rows("Issue 1: Agent speaking too fast", [49, 15, 83, 41, 73, 42, 95, 58, 15, 105, 68], INBOUND_RED),
    *issue_rows("Issue 2: Word error rate too high", [37, 14, 71, 38, 63, 38, 83, 45, 12, 82, 62], INBOUND_RED),
    *issue_rows("Issue 3: Agent response time too slow", [37, 14, 67, 32, 58, 33, 77, 43, 15, 88, 56], INBOUND_RED),
    *issue_rows("Issue 4: Agent not responding", [16, 2, 30, 11, 25, 22, 38, 10, 3, 50, 16], INBOUND_RED),
    *issue_rows("Issue 5: Agent provides incorrect information", [2, 1, 3, 0, 1, 1, 10, 11, 2, 2, 6], INBOUND_RED),
    *issue_rows("Issue 6: Agent misrepresented tool results", [1, 0, 4, 0, 3, 1, 8, 8, 1, 3, 1], INBOUND_RED),
    *issue_rows("Issue 7: Agent re-asks already answered questions", [0, 1, 3, 1, 3, 1, 7, 4, 1, 2, 3], INBOUND_RED),
    *issue_rows("Issue 8: Required lead information not captured or confirmed", [0, 2, 1, 1, 0, 0, 2, 0, 0, 1, 1], INBOUND_RED),
]

SALES_OUTBOUND = [
    ("Total calls", [1369, 9, 1906, 1158, 1879, 593, 1847, 1163, 7, 1932, 1697]),
    ("Audited calls", [89, 1, 148, 33, 190, 48, 158, 93, 1, 148, 127]),
    ("% Red Alerts", ["91%", "100%", "79%", "88%", "73%", "77%", "69%", "60%", "0%", "79%", "65%"]),
    *issue_rows("Issue 1: Word error rate too high", [71, 1, 118, 27, 154, 33, 109, 68, 1, 84, 55], OUTBOUND_RED),
    *issue_rows("Issue 2: Agent speaking too fast", [64, 1, 110, 27, 152, 35, 130, 80, 1, 104, 68], OUTBOUND_RED),
    *issue_rows("Issue 3: Agent did not follow the expected call flow", [38, 0, 72, 17, 109, 24, 96, 55, 1, 72, 51], OUTBOUND_RED),
    *issue_rows("Issue 4: Agent did not deliver the opening greeting", [22, 0, 39, 7, 62, 14, 46, 31, 1, 31, 21], OUTBOUND_RED),
    *issue_rows("Issue 5: Agent response time too slow", [20, 0, 40, 9, 70, 16, 51, 21, 1, 41, 18], OUTBOUND_RED),
    *issue_rows("Issue 6: Agent speaking too slow", [13, 0, 25, 6, 27, 6, 19, 9, 0, 8, 8], OUTBOUND_RED),
    *issue_rows("Issue 7: IVR / recorded message answered", [18, 0, 19, 5, 47, 5, 28, 24, 0, 27, 22], OUTBOUND_RED),
    *issue_rows("Issue 8: Call went to voicemail", [6, 0, 9, 0, 8, 0, 12, 4, 1, 12, 4], OUTBOUND_RED),
]

SERVICE_ROWS = ["Total calls", "Audited calls", "% Red Alerts", "Issue 1", "Issue 2", "Issue 3", "Issue 4", "Issue 5"]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
PCT_ROW_FONT = Font(italic=True, color="6B7280")


def write_section(ws, name, data_rows):
    r = ws.max_row + 1 if ws.max_row > 1 else 1
    header = [name, "Ticket", "Owner", "Status", "ETA"] + DAYS
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for label, vals in data_rows:
        row = [label, "", "", "", ""] + list(vals)
        ws.append(row)
        if label.startswith(" "):
            for c in range(1, len(row) + 1):
                ws.cell(row=ws.max_row, column=c).font = PCT_ROW_FONT
        else:
            ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["..."])
    ws.append([])


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Tracker"

    ws.append(["JIRA EPIC for Quality issues", "Created date", "Owner", "Status"])
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for ticket_url, created, owner, status in JIRA_ROWS:
        ws.append([ticket_url, created, owner, status])
    ws.append(["..."])
    ws.append([])

    write_section(ws, "Sales inbound", SALES_INBOUND)
    write_section(ws, "Sales outbound", SALES_OUTBOUND)
    write_section(ws, "Service inbound", [(label, [""] * len(DAYS)) for label in SERVICE_ROWS])
    write_section(ws, "Service outbound", [(label, [""] * len(DAYS)) for label in SERVICE_ROWS])

    ws.append([
        "Sales inbound/outbound figures pulled from observability + red-alert audit data (Aug 1-11, 2026), "
        "refetched and recomputed on 12 Aug - the previous 10 Aug figures had been captured mid-day and were "
        "far too low; the numbers here reflect the completed day. 12 Aug is excluded as it's still in progress "
        "at fetch time. Service inbound/outbound has no data pipeline yet - fill in manually."
    ])

    widths = [58, 14, 12, 40, 10] + [8] * len(DAYS)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
